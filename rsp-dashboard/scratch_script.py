import openpyxl
import sys
import json

def read_headers(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        out = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            for r in range(4, 12):
                row_data = []
                for c in range(1, 15):
                    cell = ws.cell(row=r, column=c)
                    val = str(cell.value).replace("\n", " ") if cell.value else ""
                    row_data.append(val)
                headers.append(row_data)
            
            merged = [str(m) for m in ws.merged_cells.ranges]
            
            out[sheet_name] = {
                "headers": headers,
                "merged": merged
            }
        print(json.dumps(out, indent=2))
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    read_headers(r"C:\Users\Admin\Downloads\ADAS-IER (1).xlsx")
