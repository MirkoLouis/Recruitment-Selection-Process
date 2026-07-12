import openpyxl
import sys
import json

def inspect_sheet(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        out = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cols = {}
            for col_letter, col_dim in ws.column_dimensions.items():
                cols[col_letter] = col_dim.width
            
            merged = [str(m) for m in ws.merged_cells.ranges]
            
            cells = []
            for r in range(1, min(15, ws.max_row + 1)):
                for c in range(1, min(22, ws.max_column + 1)):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None:
                        font_info = ""
                        if cell.font:
                            font_info = f"Bold: {cell.font.bold}, Italic: {cell.font.italic}, Size: {cell.font.size}"
                        align_info = ""
                        if cell.alignment:
                            align_info = f"Wrap: {cell.alignment.wrap_text}, H: {cell.alignment.horizontal}, V: {cell.alignment.vertical}"
                        
                        cells.append({
                            "coord": cell.coordinate,
                            "value": str(cell.value).replace("\n", "\\n"),
                            "font": font_info,
                            "align": align_info
                        })
            
            out[sheet_name] = {
                "columns": cols,
                "merged": merged,
                "cells": cells
            }
        
        print(json.dumps(out, indent=2))
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_sheet(r"C:\Users\Admin\Downloads\sdfsd).xlsx")
